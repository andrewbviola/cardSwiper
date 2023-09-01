import time
import os
import platform
import pandas as pd
from datetime import date
from openpyxl import load_workbook

def loadExcel(file, sheet):
        excelDoc = pd.read_excel(file, sheet_name=sheet)
        wb1 = load_workbook(file)
        ws1 = wb1[sheet]
        return excelDoc, wb1, ws1
    
def saveExcel(excelFile, newEntry, ws1, wb1):
    newEntryFrame = pd.DataFrame(newEntry)
    max_row2 = ws1.max_row
    for row_idx, row in newEntryFrame.iterrows():
        values = row.values
        for col_idx, value in enumerate(values, start=1):
            ws1.cell(row=max_row2 + row_idx + 1, column=col_idx, value=value)
    wb1.save(excelFile)
    
def timeDate():
    x = time.localtime()
    currentTime = time.strftime("%H:%M", x)
    currentDate = date.today()
    return currentTime, currentDate

def autoFillSignOut(checkForDate, excelFile, ws1, wb1, currentTime, currentDate):
    currentDate = currentDate.strftime("%Y-%m-%d")
    checkForDate['Date'] = checkForDate['Date'].astype(str)
    containsDate = checkForDate[checkForDate['Date']==currentDate]

    numRows = containsDate.shape[0]
    
    for i in range(numRows):
        currentRow = containsDate.iloc[[i]]
        if currentRow["Time Out"].to_string(index=False) == "NaN":
            indexCell = currentRow.index.item() + 2
            cell = ws1.cell(row=indexCell, column=3)
            cell.value = currentTime
            wb1.save(excelFile) 
                
def addNewUser(pid, excelFile, ws, wb):
    print("We couldn't find your information please fill out the following (Case Sensitive):")
    firstName = input("First Name: ")
    lastName = input("Last Name: ")
    email = input("Email: ")
    community = input("Community (Galileo, Hypatia, Other, N/A): ")
    year = input("Year (Freshman, Sophomore, Junior, Senior, Graduate): ")
    newData = {"PID": [pid], "First Name": [firstName], "Last Name": [lastName], "Email": [email], "Community": [community],"Year": [year]}
    saveExcel(excelFile, newData, ws, wb)
    return firstName, lastName, email, community, year

def grabData(dataFrame):
    firstName = dataFrame["First Name"].to_string(index=False)
    lastName = dataFrame["Last Name"].to_string(index=False)
    email = dataFrame["Email"].to_string(index=False)
    community = dataFrame["Community"].to_string(index=False)
    year = dataFrame["Year"].to_string(index=False)
    cnc = dataFrame["CNC"].to_string(index=False)
    lc = dataFrame["Laser Cutter"].to_string(index=False)
    sold = dataFrame["Soldering"].to_string(index=False)
    pt = dataFrame["Power Tools"].to_string(index=False)
    return firstName, lastName, email, community, year, cnc, lc, sold, pt

def firstSignIn(currentDate, currentTime, pid, firstName, lastName, email, community, year, ws1, wb1, excelFile):
    print("Signing in")
    newEntry = {"Date": [currentDate], "Time In": [currentTime], "Time Out": [""], "ID Number": [pid], "First Name": [firstName], "Last Name": [lastName], "Email": [email], "Community": [community],"Year": [year]}
    saveExcel(excelFile, newEntry, ws1, wb1)
    
def signIn(firstName, lastName, email, community, year, pid, currentTime, currentDate, cnc, lc, sold, pt, ws1, wb1, excelFile):
    print("Signing in:", firstName, lastName,"at", currentTime)
    trainedTools = {"CNC": [cnc], "Laser Cutter": [lc], "Soldering": [sold], "Power Tools": [pt]}
    dataTrained = pd.DataFrame(trainedTools)
    print(firstName, "is trained on the following tools:")
    print(dataTrained)
    newEntry = {"Date": [currentDate], "Time In": [currentTime], "Time Out": [""], "ID Number": [pid], "First Name": [firstName], "Last Name": [lastName], "Email": [email], "Community": [community],"Year": [year]}
    saveExcel(excelFile, newEntry, ws1, wb1)
    
def signOut(firstName, lastName, currentTime, indexCell, ws1, wb1, excelFile):
    print("Signing out:", firstName, lastName, "at", currentTime)
    cell = ws1.cell(row=indexCell, column=3)
    cell.value = currentTime
    wb1.save(excelFile)
    
def guiTime():
    x = time.localtime()
    currentTime = time.strftime("%I:%M %p", x)
    return currentTime

def guiDate():
    currentDate = date.today()
    currentDate = currentDate.strftime("%Y/%m/%d")
    return currentDate

def topBar(canvas, width, height):
    canvas.create_rectangle(0, 0, width, 50, fill="#861F41", outline="#861F41")
    canvas.create_text(width/2,25,font=("Roboto Mono",24),text="inVenTs Studio Sign In/Out",fill="white")

def topMessage(canvas, width, height, message, color):
    canvas.create_text(width/2,90,font=("Roboto Mono",16),text=message,fill=color)

def middleMessage(canvas, width, height, message, color):
    canvas.create_text(width/2,height/2,font=("Roboto Mono",36),text=message,fill=color)

def moveScreen(prevCanvas, nextCanvas):
    prevCanvas.forget()
    nextCanvas.pack()
    nextCanvas.focus_set()

def determineBG(canvas):
    if canvas == "mainMenuCanvas":
        return "#C64600", "white"
    elif canvas == "cardReadErrorCanvas":
        return "#F74242", "black"
    elif canvas == "fillOutCanvas" or canvas == "signInCanvas":
        return "#FFE484", "black"
    elif canvas == "confirmedSignInCanvas":
        return "#00C667", "black"

